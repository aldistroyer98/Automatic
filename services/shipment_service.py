from __future__ import annotations

import math
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

from models.shipment import (
    CATEGORY_CONSUMABLES,
    CATEGORY_CONTROLS,
    CATEGORY_ORDER,
    CATEGORY_REAGENTS,
    CATEGORY_UNCLASSIFIED,
    ShipmentAnalysis,
    ShipmentOptions,
    ShipmentPreviewRow,
    ShipmentRecord,
)


MONTH_NAMES = ("Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Set", "Oct", "Nov", "Dic")
NORMALIZED_HEADERS = (
    "Fecha",
    "Cliente",
    "ClienteCorto",
    "CodProd",
    "CodEqv",
    "ProductoLimpio",
    "Cant",
    "Año",
    "Mes",
    "Linea",
    "TipoProductoInterno",
    "ResponsableDepa",
    "Comodato",
    "Licitacion",
    "FechaMes",
    "PeriodoMes",
    "KeyClienteLineaAnio",
    "KeyTotalGeneralAnio",
    "KeyClienteProducto",
    "KeyProductoGeneral",
    "MesContempladoInicio",
    "MesContempladoFin",
    "FechaInicioProducto",
    "FechaFinProducto",
    "FechaInicioClienteProducto",
    "FechaFinClienteProducto",
    "MesInicioTotalGeneral",
    "MesFinTotalGeneral",
)

HEADER_ALIASES = {
    "fecha": ("fecha",),
    "cliente": ("cliente",),
    "cod_prod": ("codprod", "codigo producto", "cod producto"),
    "producto": ("producto", "descripcion producto"),
    "cod_eqv": ("codeqv", "codigo equivalente", "cod equivalente"),
    "cantidad": ("cant", "cantidad"),
    "anio": ("ano", "anio", "año"),
    "mes": ("mes",),
    "linea": ("linea", "lineaneg", "linea negocio"),
    "responsable": ("responsabledepa", "responsable", "responsableped"),
    "comodato": ("comodato",),
    "licitacion": ("licitacion",),
    "guia": ("guia", "guia remision", "guia de remision"),
    "pedido": ("pedido", "numero pedido", "nro pedido"),
    "lote": ("lote",),
    "expira": ("expira", "fecha expiracion", "vencimiento"),
    "registro_sanitario": ("registro sanitario", "registrosanitario"),
}

CONTROL_KEYWORDS = (
    "control", "controls", "calibrador", "calibration", "plasma de calibracion",
    "assayed", "abnormal", "ctrl",
)
REAGENT_KEYWORDS = (
    "hemosil", "thromboplastin", "recombiplastin", "synthasil", "thrombin",
    "fibrinogen", "d-dimer", "dimero", "von willebrand", "factor viii", "antigen",
    "aptt", " pt ", "drvvt", "drw", "atiii",
)
CONSUMABLE_KEYWORDS = (
    "factor diluent", "rinse", "cleaning", "clean", "diluent", "cuvettes",
    "cuvetas", "solution", "agent", "wash", "cleaner", "rotor", "rotores",
    "cups", "cup ", "printer paper",
)
CONSUMABLE_PRIORITY = (
    "rinse solution",
    "cleaning solution",
    "cleaning agent",
    "factor diluent",
    "cuvettes",
)

TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SUBTITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FILL = PatternFill("solid", fgColor="FFD966")
MONTH_FILL = PatternFill("solid", fgColor="F4B183")
INACTIVE_FILL = PatternFill("solid", fgColor="F4CCCC")
CATEGORY_FILLS = {
    CATEGORY_CONTROLS: PatternFill("solid", fgColor="DDEBF7"),
    CATEGORY_REAGENTS: PatternFill("solid", fgColor="FCE4D6"),
    CATEGORY_CONSUMABLES: PatternFill("solid", fgColor="E2F0D9"),
}
THIN_BORDER = Border(
    left=Side(style="thin", color="B7B7B7"),
    right=Side(style="thin", color="B7B7B7"),
    top=Side(style="thin", color="B7B7B7"),
    bottom=Side(style="thin", color="B7B7B7"),
)


class ShipmentValidationError(ValueError):
    pass


class ShipmentService:
    def __init__(self, reference_path: str | Path | None = None, today: date | None = None) -> None:
        self.reference_path = Path(reference_path) if reference_path else None
        self.today = today or date.today()
        self._model_order = self._load_model_order()

    @staticmethod
    def _key(value: object) -> str:
        text = unicodedata.normalize("NFKD", str(value or ""))
        return " ".join(text.encode("ascii", "ignore").decode().lower().split())

    @staticmethod
    def _text(value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return " ".join(str(value).replace("\xa0", " ").split())

    @classmethod
    def clean_product(cls, product: object, cod_prod: object = "") -> str:
        text = cls._text(product)
        code = cls._text(cod_prod)
        if code:
            text = re.sub(rf"^\s*{re.escape(code)}(?:\s+|[-:]\s*)", "", text, flags=re.IGNORECASE)
        return " ".join(text.split())

    @classmethod
    def classify_product(cls, product: object) -> str:
        normalized = f" {cls._key(product)} "
        if any(keyword in normalized for keyword in CONTROL_KEYWORDS):
            return CATEGORY_CONTROLS
        if any(keyword in normalized for keyword in CONSUMABLE_KEYWORDS):
            return CATEGORY_CONSUMABLES
        if any(keyword in normalized for keyword in REAGENT_KEYWORDS):
            return CATEGORY_REAGENTS
        return CATEGORY_UNCLASSIFIED

    @classmethod
    def short_client_name(cls, client: object) -> str:
        name = cls._text(client).upper()
        normalized = cls._key(name)
        known = (
            (("hipolito unanue",), "HNHU"),
            (("almazor", "aguinaga"), "HAAA"),
            (("almanzor", "aguinaga"), "HAAA"),
            (("iren sur",), "IREN SUR"),
            (("enfermedades neoplasicas del sur",), "IREN SUR"),
            (("salud del nino",), "INSN"),
            (("materno perinatal",), "INMP"),
        )
        for needles, abbreviation in known:
            if all(needle in normalized for needle in needles):
                return abbreviation
        words = re.findall(r"[A-Z0-9]+", unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode())
        ignored = {"DE", "DEL", "LA", "LAS", "LOS", "EL", "Y", "SAC", "SA"}
        significant = [word for word in words if word not in ignored]
        initials = "".join(word[0] for word in significant[:6])
        return initials if 2 <= len(initials) <= 10 else " ".join(significant[:3])[:31]

    @staticmethod
    def _number(value: object) -> float:
        if value is None or value == "":
            raise ValueError("cantidad vacía")
        if isinstance(value, str):
            value = value.strip().replace(",", "")
        number = float(value)
        if not math.isfinite(number):
            raise ValueError("cantidad no finita")
        return number

    @staticmethod
    def _date(value: object) -> date | datetime | None:
        if value in (None, ""):
            return None
        if isinstance(value, (date, datetime)):
            return value
        if isinstance(value, (int, float)):
            return from_excel(value)
        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        raise ValueError(f"fecha no reconocida: {text}")

    @classmethod
    def _header_map(cls, headers: Sequence[object]) -> dict[str, int]:
        normalized = {cls._key(value): index for index, value in enumerate(headers) if value is not None}
        result: dict[str, int] = {}
        for field, aliases in HEADER_ALIASES.items():
            for alias in aliases:
                if cls._key(alias) in normalized:
                    result[field] = normalized[cls._key(alias)]
                    break
        return result

    def analyze(self, source: str | Path) -> ShipmentAnalysis:
        path = Path(source)
        if not path.exists():
            raise ShipmentValidationError(f"No existe el archivo: {path}")
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise ShipmentValidationError("La base debe ser un archivo .xlsx o .xlsm")
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration as exc:
            raise ShipmentValidationError("El archivo no contiene filas") from exc
        columns = self._header_map(headers)
        required = {"cliente", "cod_prod", "producto", "cod_eqv", "cantidad", "linea"}
        missing = sorted(required - columns.keys())
        if "fecha" not in columns and not {"anio", "mes"} <= columns.keys():
            missing.append("Fecha o Año + Mes")
        if missing:
            raise ShipmentValidationError("Faltan columnas obligatorias: " + ", ".join(missing))

        records: list[ShipmentRecord] = []
        errors: list[str] = []
        rows_read = 0
        for excel_row, row in enumerate(rows, start=2):
            rows_read += 1
            try:
                client = self._text(row[columns["cliente"]])
                code = self._text(row[columns["cod_prod"]])
                product = self.clean_product(row[columns["producto"]], code)
                equivalent = self._text(row[columns["cod_eqv"]])
                line = self._text(row[columns["linea"]])
                if not client or not code or not product:
                    raise ValueError("Cliente, CodProd y Producto son obligatorios")
                shipment_date = self._date(row[columns["fecha"]]) if "fecha" in columns else None
                year_value = row[columns["anio"]] if "anio" in columns else None
                month_value = row[columns["mes"]] if "mes" in columns else None
                year = int(float(year_value)) if year_value not in (None, "") else shipment_date.year
                month = int(float(month_value)) if month_value not in (None, "") else shipment_date.month
                if not 1900 <= year <= 9999:
                    raise ValueError(f"año inválido: {year}")
                if not 1 <= month <= 12:
                    raise ValueError(f"mes inválido: {month}")
                if shipment_date is None:
                    shipment_date = date(year, month, 1)
                records.append(
                    ShipmentRecord(
                        fecha=shipment_date,
                        cliente=client,
                        cliente_corto=self.short_client_name(client),
                        cod_prod=code,
                        cod_eqv=equivalent,
                        producto=product,
                        cantidad=self._number(row[columns["cantidad"]]),
                        anio=year,
                        mes=month,
                        linea=line or "SIN LÍNEA",
                        categoria=self.classify_product(product),
                        responsable=self._optional(row, columns, "responsable"),
                        comodato=self._optional(row, columns, "comodato"),
                        licitacion=self._optional(row, columns, "licitacion"),
                        guia=self._optional(row, columns, "guia"),
                        pedido=self._optional(row, columns, "pedido"),
                        lote=self._optional(row, columns, "lote"),
                        expira=self._optional(row, columns, "expira"),
                        registro_sanitario=self._optional(
                            row, columns, "registro_sanitario"
                        ),
                    )
                )
            except (TypeError, ValueError, AttributeError) as exc:
                if len(errors) < 100:
                    errors.append(f"Fila {excel_row}: {exc}")
        workbook.close()
        if not records:
            detail = f" Primer error: {errors[0]}" if errors else ""
            raise ShipmentValidationError("No se encontraron registros válidos." + detail)
        return ShipmentAnalysis(path, tuple(records), rows_read, rows_read - len(records), tuple(errors))

    @classmethod
    def _optional(cls, row: Sequence[object], columns: dict[str, int], field: str) -> str:
        return cls._text(row[columns[field]]) if field in columns else ""

    def filter_records(
        self, records: Iterable[ShipmentRecord], options: ShipmentOptions | None = None
    ) -> list[ShipmentRecord]:
        options = options or ShipmentOptions()
        filters = (
            ("cliente", options.clients),
            ("anio", options.years),
            ("linea", options.lines),
            ("producto", options.products),
            ("categoria", options.categories),
            ("responsable", options.responsibles),
            ("comodato", options.comodatos),
            ("licitacion", options.licitaciones),
        )
        return [
            record for record in records
            if all(not selected or getattr(record, field) in selected for field, selected in filters)
        ]

    def preview(
        self, analysis: ShipmentAnalysis, options: ShipmentOptions | None = None
    ) -> list[ShipmentPreviewRow]:
        options = options or ShipmentOptions()
        filtered = self.filter_records(analysis.records, options)
        grouped: dict[tuple[str, int, str, str, str, str, str], list[ShipmentRecord]] = defaultdict(list)
        block_months: dict[tuple[str, int, str], list[int]] = defaultdict(list)
        for record in filtered:
            grouped[
                record.cliente, record.anio, record.linea, record.cod_prod,
                record.cod_eqv, record.producto, record.categoria
            ].append(record)
            if record.cantidad > 0:
                block_months[(record.cliente, record.anio, record.linea)].append(record.mes)
        result = []
        for key, rows in grouped.items():
            months = block_months[(key[0], key[1], key[2])]
            closed_month = self._last_valid_month(key[1], options)
            considered = [month for month in months if month <= closed_month]
            first_month = min(considered) if considered else 0
            last_month = max(considered) if considered else 0
            divisor = last_month - first_month + 1 if considered else 0
            total = sum(row.cantidad for row in rows if first_month <= row.mes <= last_month)
            result.append(ShipmentPreviewRow(*key, total, divisor, total / divisor if divisor else 0))
        return sorted(result, key=lambda row: (row.cliente, row.anio, row.linea, self._product_sort(row.cod_prod, row.producto)))

    def generate_report(
        self,
        analysis: ShipmentAnalysis,
        destination: str | Path,
        options: ShipmentOptions | None = None,
    ) -> Path:
        options = options or ShipmentOptions()
        records = self.filter_records(analysis.records, options)
        if not records:
            raise ShipmentValidationError("Los filtros no dejan registros para generar el reporte")
        output = Path(destination)
        if output.suffix.lower() != ".xlsx":
            output = output.with_suffix(".xlsx")
        output.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        workbook.remove(workbook.active)
        if options.create_summary:
            self._write_summary(workbook.create_sheet("Resumen"), records, options)
        self._write_report_sheet(
            workbook.create_sheet("Total General"),
            records,
            options,
            title="TOTAL GENERAL DE CLIENTES",
            client=None,
        )
        if options.create_client_sheets:
            used = set(workbook.sheetnames)
            clients = sorted({record.cliente for record in records})
            for client in clients:
                short = next(record.cliente_corto for record in records if record.cliente == client)
                sheet_name = self.safe_sheet_name(short, used)
                used.add(sheet_name)
                client_records = [record for record in records if record.cliente == client]
                self._write_report_sheet(
                    workbook.create_sheet(sheet_name), client_records, options, title=client, client=client
                )
        normalized = workbook.create_sheet("Data_Normalizada")
        self._write_normalized(normalized, records, options)
        normalized.sheet_state = "hidden" if options.hide_normalized_data else "visible"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        workbook.save(output)
        self._validate_generated_workbook(output)
        return output

    @staticmethod
    def safe_sheet_name(name: str, used: set[str] | None = None) -> str:
        used = used or set()
        base = re.sub(r"[\[\]:*?/\\]", " ", name).strip().strip("'") or "Cliente"
        base = " ".join(base.split())[:31]
        candidate = base
        suffix = 2
        while candidate.casefold() in {item.casefold() for item in used}:
            ending = f" ({suffix})"
            candidate = base[: 31 - len(ending)] + ending
            suffix += 1
        return candidate

    def _write_normalized(
        self,
        worksheet,
        records: Sequence[ShipmentRecord],
        options: ShipmentOptions,
    ) -> None:
        worksheet.append(NORMALIZED_HEADERS)
        block_periods: dict[tuple[str, str, int], tuple[int, int]] = {}
        total_periods: dict[tuple[str, int], tuple[int, int]] = {}
        product_periods: dict[str, tuple[date, date]] = {}
        client_product_periods: dict[tuple[str, str], tuple[date, date]] = {}
        valid_records = [
            record
            for record in records
            if record.cantidad > 0
            and record.mes <= self._last_valid_month(record.anio, options)
        ]
        for record in valid_records:
            block_key = (record.cliente, record.linea, record.anio)
            total_key = (record.linea, record.anio)
            block_periods[block_key] = self._extend_month_period(block_periods.get(block_key), record.mes)
            total_periods[total_key] = self._extend_month_period(total_periods.get(total_key), record.mes)
            shipment_date = self._month_date(record)
            product_periods[record.cod_prod] = self._extend_date_period(
                product_periods.get(record.cod_prod), shipment_date
            )
            client_product_key = (record.cliente, record.cod_prod)
            client_product_periods[client_product_key] = self._extend_date_period(
                client_product_periods.get(client_product_key), shipment_date
            )

        seen_blocks: set[tuple[str, str, int]] = set()
        seen_totals: set[tuple[str, int]] = set()
        seen_products: set[str] = set()
        seen_client_products: set[tuple[str, str]] = set()
        for record in records:
            block_key = (record.cliente, record.linea, record.anio)
            total_key = (record.linea, record.anio)
            client_product_key = (record.cliente, record.cod_prod)
            period = block_periods.get(block_key, (0, 0))
            total_period = total_periods.get(total_key, (0, 0))
            product_period = product_periods.get(record.cod_prod)
            client_product_period = client_product_periods.get(client_product_key)
            is_first_block = block_key not in seen_blocks
            is_first_total = total_key not in seen_totals
            is_first_product = record.cod_prod not in seen_products
            is_first_client_product = client_product_key not in seen_client_products
            worksheet.append(
                (
                    record.fecha,
                    record.cliente,
                    record.cliente_corto,
                    record.cod_prod,
                    record.cod_eqv,
                    record.producto,
                    record.cantidad,
                    record.anio,
                    record.mes,
                    record.linea,
                    record.categoria,
                    record.responsable,
                    record.comodato,
                    record.licitacion,
                    self._month_date(record),
                    record.anio * 100 + record.mes,
                    self._key_join(*block_key),
                    self._key_join(*total_key),
                    self._key_join(*client_product_key),
                    record.cod_prod,
                    period[0] if is_first_block else None,
                    period[1] if is_first_block else None,
                    product_period[0] if is_first_product and product_period else None,
                    product_period[1] if is_first_product and product_period else None,
                    client_product_period[0]
                    if is_first_client_product and client_product_period else None,
                    client_product_period[1]
                    if is_first_client_product and client_product_period else None,
                    total_period[0] if is_first_total else None,
                    total_period[1] if is_first_total else None,
                )
            )
            seen_blocks.add(block_key)
            seen_totals.add(total_key)
            seen_products.add(record.cod_prod)
            seen_client_products.add(client_product_key)
        for cell in worksheet[1]:
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
        for column in ("D", "E"):
            for cell in worksheet[column]:
                cell.number_format = "@"
        for column in ("O", "W", "X", "Y", "Z"):
            for cell in worksheet[column][1:]:
                cell.number_format = "mmm-yyyy"
        worksheet.freeze_panes = None

    @staticmethod
    def _extend_month_period(period: tuple[int, int] | None, month: int) -> tuple[int, int]:
        return (month, month) if period is None else (min(period[0], month), max(period[1], month))

    @staticmethod
    def _extend_date_period(period: tuple[date, date] | None, value: date) -> tuple[date, date]:
        return (value, value) if period is None else (min(period[0], value), max(period[1], value))

    @staticmethod
    def _month_date(record: ShipmentRecord) -> date:
        return date(record.anio, record.mes, 1)

    @staticmethod
    def _key_join(*values: object) -> str:
        return "|".join(str(value).replace("|", "/") for value in values)

    def _write_summary(self, worksheet, records: Sequence[ShipmentRecord], options: ShipmentOptions) -> None:
        worksheet.merge_cells("A1:H1")
        worksheet["A1"] = "RESUMEN GENERAL POR PRODUCTO"
        self._style_title(worksheet, 1, 8)
        headers = ("CodProd", "CodEqv", "Producto", "Fecha Inicio", "Fecha Fin", "Total", "Mes", "Prod")
        for column, header in enumerate(headers, 1):
            worksheet.cell(3, column, header)
        self._style_header(worksheet, 3, 8)

        row = 4
        products = self._distinct_products(records)
        cutoff_month = self.today.month if options.exclude_current_month else self.today.month + 1
        cutoff_year = self.today.year
        if cutoff_month == 13:
            cutoff_year += 1
            cutoff_month = 1
        cutoff = f"DATE({cutoff_year},{cutoff_month},1)"
        for code, equivalent, product, category in products:
            worksheet.cell(row, 1, code).number_format = "@"
            worksheet.cell(row, 2, equivalent).number_format = "@"
            worksheet.cell(row, 3, product)
            worksheet.cell(row, 4, f'=IFERROR(SUMIFS(\'Data_Normalizada\'!$W:$W,\'Data_Normalizada\'!$T:$T,A{row}),0)')
            worksheet.cell(row, 5, f'=IFERROR(SUMIFS(\'Data_Normalizada\'!$X:$X,\'Data_Normalizada\'!$T:$T,A{row}),0)')
            worksheet.cell(row, 6, f'=SUMIFS(\'Data_Normalizada\'!$G:$G,\'Data_Normalizada\'!$D:$D,A{row},\'Data_Normalizada\'!$O:$O,"<"&{cutoff})')
            worksheet.cell(row, 7, f'=IF(OR(D{row}=0,E{row}=0),0,(YEAR(E{row})-YEAR(D{row}))*12+MONTH(E{row})-MONTH(D{row})+1)')
            worksheet.cell(row, 8, f'=IF(G{row}=0,0,F{row}/G{row})')
            worksheet.cell(row, 4).number_format = "mmm-yyyy"
            worksheet.cell(row, 5).number_format = "mmm-yyyy"
            self._style_summary_row(worksheet, row, category, 8)
            row += 1

        row += 2
        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        worksheet.cell(row, 1, "RESUMEN POR CLIENTE Y PRODUCTO")
        self._style_title(worksheet, row, 8)
        row += 1
        for client in sorted({record.cliente for record in records}):
            worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            worksheet.cell(row, 1, client.upper())
            self._style_title(worksheet, row, 8)
            row += 1
            for column, header in enumerate(headers, 1):
                worksheet.cell(row, column, header)
            self._style_header(worksheet, row, 8)
            row += 1
            client_records = [record for record in records if record.cliente == client]
            for code, equivalent, product, category in self._distinct_products(client_records):
                key = self._key_join(client, code)
                worksheet.cell(row, 1, code).number_format = "@"
                worksheet.cell(row, 2, equivalent).number_format = "@"
                worksheet.cell(row, 3, product)
                worksheet.cell(row, 4, f'=IFERROR(SUMIFS(\'Data_Normalizada\'!$Y:$Y,\'Data_Normalizada\'!$S:$S,{self._excel_string(key)}),0)')
                worksheet.cell(row, 5, f'=IFERROR(SUMIFS(\'Data_Normalizada\'!$Z:$Z,\'Data_Normalizada\'!$S:$S,{self._excel_string(key)}),0)')
                worksheet.cell(row, 6, f'=SUMIFS(\'Data_Normalizada\'!$G:$G,\'Data_Normalizada\'!$B:$B,{self._excel_string(client)},\'Data_Normalizada\'!$D:$D,A{row},\'Data_Normalizada\'!$O:$O,"<"&{cutoff})')
                worksheet.cell(row, 7, f'=IF(OR(D{row}=0,E{row}=0),0,(YEAR(E{row})-YEAR(D{row}))*12+MONTH(E{row})-MONTH(D{row})+1)')
                worksheet.cell(row, 8, f'=IF(G{row}=0,0,F{row}/G{row})')
                worksheet.cell(row, 4).number_format = "mmm-yyyy"
                worksheet.cell(row, 5).number_format = "mmm-yyyy"
                self._style_summary_row(worksheet, row, category, 8)
                row += 1
            row += 1

        for column, width in enumerate((12, 12, 36, 12, 12, 5, 5, 5), 1):
            worksheet.column_dimensions[get_column_letter(column)].width = width
        worksheet.freeze_panes = None

    def _write_report_sheet(
        self,
        worksheet,
        records: Sequence[ShipmentRecord],
        options: ShipmentOptions,
        title: str,
        client: str | None,
    ) -> None:
        row = 1
        for line in sorted({record.linea for record in records}):
            line_records = [record for record in records if record.linea == line]
            for year in sorted({record.anio for record in line_records}):
                block = [record for record in line_records if record.anio == year]
                products = self._distinct_products(block)
                worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=18)
                worksheet.cell(row, 1, title.upper())
                self._style_title(worksheet, row, 18)
                row += 1
                worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                worksheet.merge_cells(start_row=row, start_column=4, end_row=row, end_column=18)
                worksheet.cell(row, 1, line.upper())
                worksheet.cell(row, 4, year)
                self._style_subtitle(worksheet, row)
                row += 1
                headers = ("CodProd", "CodEqv", "Producto", *MONTH_NAMES, "Total", "Mes", "Prod")
                for column, header in enumerate(headers, 1):
                    worksheet.cell(row, column, header)
                self._style_header(worksheet, row, 18)
                row += 1
                first_month, last_month = self._block_period(block, year, options)
                for code, equivalent, product, category in products:
                    worksheet.cell(row, 1, code).number_format = "@"
                    worksheet.cell(row, 2, equivalent).number_format = "@"
                    worksheet.cell(row, 3, product)
                    for month in range(1, 13):
                        cell = worksheet.cell(row, month + 3)
                        if month < first_month or month > last_month:
                            cell.value = '=""'
                            cell.fill = INACTIVE_FILL
                        else:
                            criteria = [
                                "'Data_Normalizada'!$D:$D", self._excel_string(code),
                                "'Data_Normalizada'!$H:$H", str(year),
                                "'Data_Normalizada'!$I:$I", str(month),
                                "'Data_Normalizada'!$J:$J", self._excel_string(line),
                            ]
                            if client:
                                criteria.extend(
                                    ("'Data_Normalizada'!$B:$B", self._excel_string(client))
                                )
                            pairs = ",".join(f"{criteria[index]},{criteria[index + 1]}" for index in range(0, len(criteria), 2))
                            cell.value = f"=SUMIFS('Data_Normalizada'!$G:$G,{pairs})"
                            cell.number_format = "0"
                    worksheet.cell(row, 16, f"=SUM(D{row}:O{row})")
                    worksheet.cell(
                        row, 17,
                        self._block_months_formula(year, line, client, options),
                    )
                    worksheet.cell(row, 18, f'=IF(Q{row}=0,0,P{row}/Q{row})')
                    worksheet.cell(row, 18).number_format = "0.0"
                    if options.use_category_colors and category in CATEGORY_FILLS:
                        for column in (3, 16, 17, 18):
                            worksheet.cell(row, column).fill = CATEGORY_FILLS[category]
                    for column in range(1, 19):
                        cell = worksheet.cell(row, column)
                        cell.border = THIN_BORDER
                        cell.alignment = Alignment(
                            horizontal="center",
                            vertical="center",
                            wrap_text=False,
                        )
                    worksheet.row_dimensions[row].height = 20
                    row += 1
                row += 1
        worksheet.freeze_panes = None
        widths = {1: 12, 2: 12, 3: 36, 16: 5, 17: 5, 18: 5}
        for column in range(1, 19):
            worksheet.column_dimensions[get_column_letter(column)].width = widths.get(column, 4)

    def _block_period(
        self,
        records: Sequence[ShipmentRecord],
        year: int,
        options: ShipmentOptions,
    ) -> tuple[int, int]:
        closed_month = self._last_valid_month(year, options)
        months = [
            record.mes
            for record in records
            if record.cantidad > 0 and record.mes <= closed_month
        ]
        return (min(months), max(months)) if months else (0, 0)

    def _block_months_formula(
        self,
        year: int,
        line: str,
        client: str | None,
        options: ShipmentOptions,
    ) -> str:
        if client:
            key_column = "$Q:$Q"
            start_column = "$U:$U"
            end_column = "$V:$V"
            key = self._key_join(client, line, year)
        else:
            key_column = "$R:$R"
            start_column = "$AA:$AA"
            end_column = "$AB:$AB"
            key = self._key_join(line, year)
        start = (
            f"SUMIFS('Data_Normalizada'!{start_column},"
            f"'Data_Normalizada'!{key_column},{self._excel_string(key)})"
        )
        end = (
            f"SUMIFS('Data_Normalizada'!{end_column},"
            f"'Data_Normalizada'!{key_column},{self._excel_string(key)})"
        )
        return f"=IFERROR(IF({end}=0,0,{end}-{start}+1),0)"

    @staticmethod
    def _style_summary_row(worksheet, row: int, category: str, last_column: int) -> None:
        for column in range(1, last_column + 1):
            cell = worksheet.cell(row, column)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=False,
            )
        worksheet.row_dimensions[row].height = 20
        colored_columns = (3, 6, 7, 8) if last_column == 8 else (4, 7, 8, 9)
        for column in colored_columns:
            if category in CATEGORY_FILLS:
                worksheet.cell(row, column).fill = CATEGORY_FILLS[category]
        worksheet.cell(row, last_column).number_format = "0.0"

    @staticmethod
    def _validate_generated_workbook(path: Path) -> None:
        workbook = load_workbook(path, data_only=False)
        expected = {"Resumen", "Total General", "Data_Normalizada"}
        unsupported_functions = (
            "MINIFS(",
            "MAXIFS(",
            "DATEDIF(",
            "EOMONTH(",
            "SI(",
            "SUMAR.SI",
            "CONTAR(",
            "PROMEDIO(",
        )
        if not expected <= set(workbook.sheetnames):
            missing = ", ".join(sorted(expected - set(workbook.sheetnames)))
            workbook.close()
            raise ShipmentValidationError(f"Faltan hojas requeridas: {missing}")
        for worksheet in workbook.worksheets:
            if worksheet.tables:
                workbook.close()
                raise ShipmentValidationError(
                    f"La hoja {worksheet.title} contiene tablas estructuradas no permitidas"
                )
            if worksheet.auto_filter.ref:
                workbook.close()
                raise ShipmentValidationError(
                    f"La hoja {worksheet.title} contiene un autofiltro no permitido"
                )
            for row in worksheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        upper_formula = cell.value.upper()
                        if "#REF!" in upper_formula:
                            workbook.close()
                            raise ShipmentValidationError(
                                f"Fórmula inválida en {worksheet.title}!{cell.coordinate}"
                            )
                        invalid = next(
                            (
                                function
                                for function in unsupported_functions
                                if function in upper_formula
                            ),
                            None,
                        )
                        if invalid:
                            workbook.close()
                            raise ShipmentValidationError(
                                f"Función no compatible {invalid} en "
                                f"{worksheet.title}!{cell.coordinate}"
                            )
        workbook.save(path)
        workbook.close()

    def _distinct_products(self, records: Sequence[ShipmentRecord]) -> list[tuple[str, str, str, str]]:
        products: dict[str, tuple[str, str, str, str]] = {}
        for record in records:
            products.setdefault(
                record.cod_prod,
                (record.cod_prod, record.cod_eqv, record.producto, record.categoria),
            )
        return sorted(products.values(), key=lambda item: self._product_sort(item[0], item[2]))

    def _product_sort(self, code: str, product: str) -> tuple[int, int, int, str, str]:
        category = self.classify_product(product)
        consumable_position = self._consumable_position(product) if category == CATEGORY_CONSUMABLES else 0
        model_position = self._model_order.get(code, 1_000_000)
        return (
            CATEGORY_ORDER[category],
            consumable_position,
            model_position,
            code.casefold(),
            product.casefold(),
        )

    @classmethod
    def _consumable_position(cls, product: str) -> int:
        normalized = cls._key(product)
        for index, keyword in enumerate(CONSUMABLE_PRIORITY):
            if keyword in normalized:
                return index
        return len(CONSUMABLE_PRIORITY)

    def _load_model_order(self) -> dict[str, int]:
        if not self.reference_path or not self.reference_path.exists():
            return {}
        workbook = load_workbook(self.reference_path, read_only=True, data_only=False)
        result: dict[str, int] = {}
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                code = self._text(row[0] if row else "")
                if re.match(r"^[A-Z]{2,}\d", code) and code not in result:
                    result[code] = len(result)
        workbook.close()
        return result

    def _last_valid_month(self, year: int, options: ShipmentOptions) -> int:
        if year < self.today.year:
            return 12
        if year > self.today.year:
            return 0
        return max(0, self.today.month - 1) if options.exclude_current_month else self.today.month

    @staticmethod
    def _excel_string(value: object) -> str:
        return '"' + str(value).replace('"', '""') + '"'

    @staticmethod
    def _style_title(worksheet, row: int, last_column: int) -> None:
        cell = worksheet.cell(row, 1)
        cell.fill = TITLE_FILL
        cell.font = Font(color="FFFFFF", bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[row].height = 24
        for column in range(1, last_column + 1):
            worksheet.cell(row, column).fill = TITLE_FILL

    @staticmethod
    def _style_subtitle(worksheet, row: int) -> None:
        for column in range(1, 19):
            cell = worksheet.cell(row, column)
            cell.fill = SUBTITLE_FILL
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def _style_header(worksheet, row: int, last_column: int) -> None:
        for column in range(1, last_column + 1):
            cell = worksheet.cell(row, column)
            cell.fill = MONTH_FILL if 4 <= column <= 15 else HEADER_FILL
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
