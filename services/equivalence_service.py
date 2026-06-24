from __future__ import annotations

import csv
import json
import math
import re
import shutil
import unicodedata
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models.equivalence import EquivalenceResult, EquivalenceState, ImportPreviewRow, ReagentProduct, TenderTest
from services.category_manager import CategoryManager
from services.excel_reader import sanitize_tabular_rows


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
WARNING_FILL = PatternFill("solid", fgColor="FCE4D6")
ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
THIN_BORDER = Border(
    left=Side(style="thin", color="B7B7B7"),
    right=Side(style="thin", color="B7B7B7"),
    top=Side(style="thin", color="B7B7B7"),
    bottom=Side(style="thin", color="B7B7B7"),
)


def normalize_description(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return " ".join(text.encode("ascii", "ignore").decode().casefold().split())


class EquivalenceService:
    def __init__(self, config_path: str | Path) -> None:
        self.config_path = Path(config_path)

    def load(self) -> EquivalenceState:
        if not self.config_path.exists():
            return EquivalenceState()
        try:
            data = json.loads(self.config_path.read_text(encoding="utf-8"))
            products = [
                ReagentProduct(
                    cod_prod=str(item.get("cod_prod", "")).strip(),
                    cod_eqv=str(item.get("cod_eqv", "")).strip(),
                    product=str(item.get("product", "")).strip(),
                    det_rvo=self._number(item.get("det_rvo", 0), default=0),
                    category=CategoryManager.internal_name(
                        str(item.get("category", "")).strip()
                    ),
                    order=int(item.get("order", index)),
                )
                for index, item in enumerate(data.get("products", []))
                if isinstance(item, dict)
            ]
            equivalences = {
                normalize_description(key): [str(value).strip() for value in values if str(value).strip()]
                for key, values in data.get("equivalences", {}).items()
                if isinstance(values, list)
            }
            settings = data.get("settings", {})
            return EquivalenceState(products, equivalences, settings if isinstance(settings, dict) else {})
        except (OSError, ValueError, TypeError, json.JSONDecodeError):
            return EquivalenceState()

    def save(self, state: EquivalenceState) -> None:
        self.config_path.parent.mkdir(parents=True, exist_ok=True)
        data = {
            "version": 1,
            "products": [
                {
                    "cod_prod": product.cod_prod,
                    "cod_eqv": product.cod_eqv,
                    "product": product.product,
                    "det_rvo": product.det_rvo,
                    "category": product.category,
                    "order": index,
                }
                for index, product in enumerate(self.sorted_products(state.products))
            ],
            "equivalences": {
                normalize_description(key): list(dict.fromkeys(values))
                for key, values in state.equivalences.items()
            },
            "settings": state.settings,
        }
        self.config_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    def parse_clipboard_text(self, text: str) -> list[TenderTest]:
        rows: list[TenderTest] = []
        for raw_line in text.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            parsed = self._parse_tender_line(line)
            if parsed is not None:
                rows.append(parsed)
        return rows

    def parse_import_text(self, text: str) -> list[ImportPreviewRow]:
        rows: list[ImportPreviewRow] = []
        for raw_line in text.splitlines():
            line = raw_line.strip()
            if not line or self._looks_like_header([line]):
                continue
            parsed = self._parse_import_line(line)
            if parsed is not None:
                rows.append(parsed)
        return rows

    def extract_import_preview(self, path: str | Path) -> tuple[list[ImportPreviewRow], str]:
        source = Path(path)
        suffix = source.suffix.lower()
        if suffix == ".pdf":
            text = self._extract_pdf_text(source)
            rows = self.parse_import_text(text)
            if self._has_useful_preview(rows):
                return rows, text
            text = self._ocr_pdf(source)
            return self.parse_import_text(text), text
        if suffix in {".png", ".jpg", ".jpeg", ".bmp"}:
            text = self._ocr_image_file(source)
            return self.parse_import_text(text), text
        raise ValueError("Formato no soportado. Usa PNG, JPG, JPEG o PDF.")

    def export_import_review(self, path: str | Path, rows: Iterable[ImportPreviewRow]) -> Path:
        output = Path(path)
        if output.suffix.lower() == ".csv":
            output.parent.mkdir(parents=True, exist_ok=True)
            with output.open("w", encoding="utf-8-sig", newline="") as file:
                writer = csv.writer(file)
                writer.writerow(["Código SAP", "Descripción", "Cantidad", "Estado", "Observación"])
                for row in rows:
                    writer.writerow([row.sap_code, row.description, row.quantity, row.status, row.observation])
            return output

        if output.suffix.lower() != ".xlsx":
            output = output.with_suffix(".xlsx")
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Revision"
        headers = ("Código SAP", "Descripción", "Cantidad", "Estado", "Observación")
        self._write_headers(sheet, 1, headers)
        critical = {"Fila incompleta", "Revisar código", "Revisar cantidad"}
        for row_index, row in enumerate(rows, 2):
            fill = ERROR_FILL if row.status in critical else (WARNING_FILL if row.status != "OK" else None)
            values = (row.sap_code, row.description, row.quantity, row.status, row.observation)
            for column, value in enumerate(values, 1):
                cell = sheet.cell(row_index, column, value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="left" if column in (2, 5) else "center")
                if fill is not None:
                    cell.fill = fill
        for column, width in {1: 16, 2: 48, 3: 14, 4: 20, 5: 42}.items():
            sheet.column_dimensions[get_column_letter(column)].width = width
        workbook.save(output)
        return output

    def _ocr_image_file(self, path: Path) -> str:
        try:
            from PIL import Image  # type: ignore
        except ImportError as exc:
            raise RuntimeError(
                "OCR no disponible. Instala Pillow y pytesseract, y verifica Tesseract OCR local."
            ) from exc
        with Image.open(path) as image:
            return self._ocr_image(image)

    def _ocr_pdf(self, path: Path) -> str:
        try:
            import fitz  # type: ignore
            from PIL import Image  # type: ignore
        except ImportError as exc:
            raise RuntimeError("PDF escaneado no disponible. Instala PyMuPDF y Pillow.") from exc

        texts = []
        with fitz.open(str(path)) as document:
            for page in document:
                pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
                image = Image.frombytes("RGB", (pixmap.width, pixmap.height), pixmap.samples)
                texts.append(self._ocr_image(image))
        return "\n".join(texts)

    @staticmethod
    def _extract_pdf_text(path: Path) -> str:
        try:
            import fitz  # type: ignore
        except ImportError as exc:
            raise RuntimeError("Importacion PDF no disponible. Instala PyMuPDF.") from exc

        with fitz.open(str(path)) as document:
            return "\n".join(page.get_text("text") for page in document)

    def _ocr_image(self, image) -> str:
        try:
            import pytesseract  # type: ignore
            from PIL import ImageEnhance, ImageFilter, ImageOps  # type: ignore
        except ImportError as exc:
            raise RuntimeError(
                "OCR no disponible. Instala Pillow y pytesseract, y verifica Tesseract OCR local."
            ) from exc

        self._ensure_tesseract_available(pytesseract)
        image = ImageOps.grayscale(image)
        image = ImageOps.autocontrast(image)
        scale = 2 if max(image.size) < 2600 else 1
        if scale > 1:
            image = image.resize((image.width * scale, image.height * scale))
        image = ImageEnhance.Contrast(image).enhance(2.0)
        image = image.filter(ImageFilter.MedianFilter(size=3))
        image = image.point(lambda value: 255 if value > 165 else 0)
        return pytesseract.image_to_string(image, config="--oem 3 --psm 6")

    @staticmethod
    def _ensure_tesseract_available(pytesseract) -> None:
        if not shutil.which("tesseract"):
            candidates = (
                Path("C:/Program Files/Tesseract-OCR/tesseract.exe"),
                Path("C:/Program Files (x86)/Tesseract-OCR/tesseract.exe"),
            )
            match = next((candidate for candidate in candidates if candidate.exists()), None)
            if match is not None:
                pytesseract.pytesseract.tesseract_cmd = str(match)
        try:
            pytesseract.get_tesseract_version()
        except Exception as exc:
            raise RuntimeError(
                "Tesseract OCR no esta disponible. Instala Tesseract local o usa pegado manual, Excel/CSV o edicion manual."
            ) from exc

    def load_tender_file(self, path: str | Path) -> list[TenderTest]:
        source = Path(path)
        if source.suffix.lower() == ".csv":
            return self._load_tender_csv(source)
        if source.suffix.lower() in {".xlsx", ".xlsm"}:
            return self._load_tender_excel(source)
        raise ValueError("Formato no soportado. Usa CSV o Excel.")

    def load_products_file(self, path: str | Path) -> list[ReagentProduct]:
        source = Path(path)
        if source.suffix.lower() == ".csv":
            return self._load_products_csv(source)
        if source.suffix.lower() in {".xlsx", ".xlsm"}:
            return self._load_products_excel(source)
        raise ValueError("Formato no soportado. Usa CSV o Excel.")

    def export_products_excel(
        self,
        path: str | Path,
        products: Iterable[ReagentProduct],
    ) -> Path:
        output = Path(path)
        if output.suffix.lower() != ".xlsx":
            output = output.with_suffix(".xlsx")
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Reactivos"
        headers = ("CodProd", "CodEqv", "Producto", "DET RVO", "Categoría", "Orden")
        self._write_headers(sheet, 1, headers)
        for row, product in enumerate(self.sorted_products(products), 2):
            values = (
                product.cod_prod,
                product.cod_eqv,
                product.product,
                product.det_rvo,
                product.category,
                product.order + 1,
            )
            for column, value in enumerate(values, 1):
                cell = sheet.cell(row, column, value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="left" if column == 3 else "center")
                if column == 4:
                    cell.number_format = "#,##0.######"
        for column, width in {1: 18, 2: 18, 3: 48, 4: 14, 5: 24, 6: 10}.items():
            sheet.column_dimensions[get_column_letter(column)].width = width
        workbook.save(output)
        return output

    def calculate(
        self,
        tests: Iterable[TenderTest],
        state: EquivalenceState,
        period_months: int,
        period_type: str,
    ) -> tuple[list[EquivalenceResult], list[str]]:
        products_by_key = {product.key: product for product in state.products if product.key}
        results: list[EquivalenceResult] = []
        warnings: list[str] = []
        seen_tests: set[tuple[str, str]] = set()

        for test in tests:
            test_key = (test.sap_code.strip(), normalize_description(test.description))
            if test_key in seen_tests:
                warnings.append(f"Duplicado: {test.sap_code} | {test.description}")
            seen_tests.add(test_key)
            if not test.description.strip() or test.quantity <= 0:
                warnings.append(f"Cantidad inválida o descripción vacía: {test.sap_code} | {test.description}")
                continue
            det_oc = self.det_oc(test.quantity, period_months, period_type)
            product_keys = state.equivalences.get(normalize_description(test.description), [])
            if not product_keys:
                warnings.append(f"Sin equivalencia: {test.sap_code} | {test.description}")
                continue
            for product_key in product_keys:
                product = products_by_key.get(product_key)
                if product is None:
                    warnings.append(f"Producto no encontrado para {test.description}: {product_key}")
                    continue
                if product.det_rvo <= 0:
                    warnings.append(f"Producto sin DET RVO: {product.cod_prod} | {product.product}")
                    results.append(self._result(test, product, det_oc, 0, 0, "Falta DET RVO"))
                    continue
                quantity = math.ceil(det_oc / product.det_rvo)
                det_env = quantity * product.det_rvo
                results.append(self._result(test, product, det_oc, det_env, quantity))
        return results, warnings

    @staticmethod
    def det_oc(quantity: float, period_months: int, period_type: str) -> float:
        months = max(1, int(period_months or 1))
        if period_type == "mensual":
            return quantity * months
        if period_type == "bimensual":
            return quantity * math.ceil(months / 2)
        return quantity

    @staticmethod
    def sorted_products(products: Iterable[ReagentProduct]) -> list[ReagentProduct]:
        return sorted(products, key=lambda item: (item.category.casefold(), item.order, item.product.casefold()))

    def export_excel(
        self,
        path: str | Path,
        tests: Iterable[TenderTest],
        results: Iterable[EquivalenceResult],
        warnings: Iterable[str],
        *,
        customer: str = "",
        line: str = "",
        equipment: str = "",
        period_label: str = "",
        categories: Iterable[dict] | None = None,
    ) -> Path:
        output = Path(path)
        if output.suffix.lower() != ".xlsx":
            output = output.with_suffix(".xlsx")
        output.parent.mkdir(parents=True, exist_ok=True)

        test_list = list(tests)
        result_list = list(results)
        warning_list = list(warnings)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Equivalencia"

        row = 1
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        sheet.cell(row, 1, "TOTAL GENERAL DE CLIENTES")
        self._style_title(sheet, row, 9)
        row += 2
        sheet.cell(row, 1, "Hospital / Cliente")
        sheet.cell(row, 2, customer)
        sheet.cell(row, 5, "Equipo")
        sheet.cell(row, 6, equipment)
        row += 1
        sheet.cell(row, 1, "Linea")
        sheet.cell(row, 2, line)
        sheet.cell(row, 5, "Periodo")
        sheet.cell(row, 6, period_label)
        row += 2

        headers = ("Codigo SAP", "Descripcion", "CodProd", "CodEqv", "Producto", "DET RVO", "DET OC", "DET ENV", "CANT")
        row = self._write_grouped_results(sheet, row, headers, result_list, categories)

        pending = self.pending_tests(test_list, result_list)
        if pending:
            row += 1
            row = self._write_table(
                sheet,
                row,
                "PENDIENTES DE HOMOLOGACION",
                ("Codigo SAP", "Descripcion", "Cantidad"),
                ((test.sap_code, test.description, test.quantity) for test in pending),
                warning=True,
            )
        if warning_list:
            row += 1
            self._write_table(sheet, row, "Alertas", ("Detalle",), ((warning,) for warning in warning_list), warning=True)

        widths = {1: 16, 2: 38, 3: 16, 4: 16, 5: 42, 6: 12, 7: 12, 8: 12, 9: 10}
        for column, width in widths.items():
            sheet.column_dimensions[get_column_letter(column)].width = width
        workbook.save(output)
        return output

    @staticmethod
    def pending_tests(tests: Iterable[TenderTest], results: Iterable[EquivalenceResult]) -> list[TenderTest]:
        resolved = {
            (result.sap_code.strip(), normalize_description(result.test_description))
            for result in results
            if not result.warning
        }
        return [
            test for test in tests
            if (test.sap_code.strip(), normalize_description(test.description)) not in resolved
        ]

    def _load_tender_csv(self, path: Path) -> list[TenderTest]:
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            rows = list(csv.reader(file))
        if not rows:
            return []

        header = [normalize_description(value).replace(" ", "_") for value in rows[0]]
        aliases = {
            "sap_code": lambda value: value in {"codigo_sap", "codigo", "cod_sap"},
            "description": lambda value: value.startswith(("descrip", "description")),
            "quantity": lambda value: value.startswith(("cantidad", "quantity", "cant")),
        }
        columns = {
            field: next((index for index, value in enumerate(header) if matches(value)), -1)
            for field, matches in aliases.items()
        }
        if all(index >= 0 for index in columns.values()):
            result = []
            for row in rows[1:]:
                values = [
                    self._text(row[columns[field]]) if columns[field] < len(row) else ""
                    for field in ("sap_code", "description", "quantity")
                ]
                if any(values):
                    result.append(TenderTest(values[0], values[1], self._number(values[2], default=0)))
            return result
        return self._rows_to_tests(rows)

    def _load_tender_excel(self, path: Path) -> list[TenderTest]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            rows = list(workbook.active.iter_rows(values_only=True))
            return self._rows_to_tests(rows)
        finally:
            workbook.close()

    def _load_products_csv(self, path: Path) -> list[ReagentProduct]:
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            rows = list(csv.reader(file))
        return self._rows_to_products(rows)

    def _load_products_excel(self, path: Path) -> list[ReagentProduct]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            rows = list(workbook.active.iter_rows(values_only=True))
            return self._rows_to_products(rows)
        finally:
            workbook.close()

    def _rows_to_tests(self, rows) -> list[TenderTest]:
        rows = sanitize_tabular_rows(rows)
        result = []
        for row in rows:
            values = [self._text(value) for value in row[:3]]
            if len(values) < 3 or self._looks_like_header(values):
                continue
            if not any(values):
                continue
            result.append(TenderTest(values[0], values[1], self._number(values[2], default=0)))
        return result

    def _rows_to_products(self, rows) -> list[ReagentProduct]:
        rows = sanitize_tabular_rows(rows)
        if not rows:
            return []
        header = [normalize_description(value).replace(" ", "_") for value in rows[0]]
        aliases = {
            "cod_prod": lambda value: value in {"codprod", "cod_prod", "codigo_producto"},
            "cod_eqv": lambda value: value in {"codeqv", "cod_eqv", "codigo_equivalente"},
            "product": lambda value: value in {"producto", "product", "descripcion"},
            "det_rvo": lambda value: value.replace("_", "") == "detrvo",
            "category": lambda value: value.startswith("categoria"),
            "order": lambda value: value in {"orden", "order"},
        }
        columns = {
            field: next((index for index, value in enumerate(header) if matches(value)), -1)
            for field, matches in aliases.items()
        }
        required = ("cod_prod", "cod_eqv", "product", "det_rvo")
        if all(columns[field] >= 0 for field in required):
            result = []
            for source_row in rows[1:]:
                def value(field: str) -> str:
                    column = columns[field]
                    return self._text(source_row[column]) if 0 <= column < len(source_row) else ""
                if not any(value(field) for field in required):
                    continue
                det_rvo = self._number(value("det_rvo"), default=-1)
                if det_rvo < 0:
                    raise ValueError("DET RVO debe ser numérico y no negativo.")
                order_text = value("order")
                order = max(0, int(self._number(order_text, default=len(result) + 1)) - 1)
                result.append(ReagentProduct(
                    cod_prod=value("cod_prod"),
                    cod_eqv=value("cod_eqv"),
                    product=value("product"),
                    det_rvo=det_rvo,
                    category=CategoryManager.internal_name(value("category"))
                    or "Sin Categoría",
                    order=order,
                ))
            return result

        recognized_headers = sum(index >= 0 for index in columns.values())
        if recognized_headers:
            missing = [field for field in required if columns[field] < 0]
            raise ValueError(
                "Faltan encabezados obligatorios: " + ", ".join(missing)
            )

        result = []
        for row in rows:
            values = [self._text(value) for value in row[:6]]
            if len(values) < 4 or self._looks_like_header(values):
                continue
            if not any(values):
                continue
            det_rvo = self._number(values[3], default=-1)
            if det_rvo < 0:
                raise ValueError("DET RVO debe ser numérico y no negativo.")
            result.append(
                ReagentProduct(
                    cod_prod=values[0],
                    cod_eqv=values[1],
                    product=values[2],
                    det_rvo=det_rvo,
                    category=(
                        CategoryManager.internal_name(values[4])
                        if len(values) > 4 and values[4]
                        else "Sin Categoría"
                    ),
                    order=len(result),
                )
            )
        return result

    def _parse_tender_line(self, line: str) -> TenderTest | None:
        if self._looks_like_header([line]):
            return None
        preview = self._parse_import_line(line)
        if preview is not None and preview.sap_code and preview.description and preview.quantity > 0:
            return TenderTest(preview.sap_code, preview.description, preview.quantity)
        match = re.match(r"^\s*(\d{6,})\s+(.+?)\s+([\d.,]+)\s*$", line)
        if match:
            return TenderTest(match.group(1), match.group(2).strip(" |-"), self._number(match.group(3), default=0))
        parts = [part.strip() for part in re.split(r"\t|\s{2,}|\|", line) if part.strip()]
        if len(parts) >= 3:
            return TenderTest(parts[0], parts[1], self._number(parts[2], default=0))
        return None

    def _parse_import_line(self, line: str) -> ImportPreviewRow | None:
        code_match = self._find_sap_code(line)
        if code_match is None:
            if self._looks_like_data_line(line):
                return self.validate_import_row("", self._clean_import_description(line), 0, line)
            return None

        sap_code = code_match.group(1)
        tail = line[code_match.end():]
        quantity_match = self._last_quantity_match(tail)
        quantity = self._number(quantity_match.group(1), default=0) if quantity_match is not None else 0
        description_source = tail[:quantity_match.start()] if quantity_match is not None else tail
        description = self._clean_import_description(description_source)
        return self.validate_import_row(sap_code, description, quantity, line)

    def validate_import_row(
        self,
        sap_code: object,
        description: object,
        quantity: object,
        source_line: str = "",
    ) -> ImportPreviewRow:
        code = self._text(sap_code)
        desc = self._clean_import_description(self._text(description))
        amount = self._number(quantity, default=0)
        observations: list[str] = []

        if not code or not desc:
            status = "Fila incompleta"
            if not code:
                observations.append("Falta codigo SAP")
            if not desc:
                observations.append("Falta descripcion")
        elif not re.fullmatch(r"\d{8}", code) or not code.startswith("30"):
            status = "Revisar código"
            observations.append("El código debe ser numérico, de 8 dígitos y empezar con 30")
        elif amount <= 0:
            status = "Revisar cantidad"
            observations.append("Cantidad vacia o menor/igual a cero")
        elif self._looks_like_weak_description(desc):
            status = "Descripción dudosa"
            observations.append("Descripción muy corta o contaminada por columnas")
        else:
            status = "OK"

        return ImportPreviewRow(code, desc, amount, status, "; ".join(observations), source_line)

    @staticmethod
    def _find_sap_code(line: str):
        candidates = list(re.finditer(r"(?<!\d)(\d{8})(?!\d)", line))
        if candidates:
            return next((match for match in candidates if match.group(1).startswith("30")), candidates[0])
        return None

    @staticmethod
    def _last_quantity_match(text: str):
        pattern = r"(?<![\w.])(\d{1,3}(?:,\d{3})+|\d+(?:[.,]\d+)?)(?![\w])"
        matches = list(re.finditer(pattern, text))
        return matches[-1] if matches else None

    @staticmethod
    def _clean_import_description(text: str) -> str:
        cleaned = re.sub(r"[\t|]+", " ", str(text or ""))
        cleaned = re.sub(r"\s+", " ", cleaned).strip(" -:;,.")
        tokens = cleaned.split()
        unit_tokens = {"PBA", "UND", "UN", "EA", "KIT", "TEST", "PRUEBA", "PRUEBAS"}
        while tokens and tokens[-1].upper().strip(".,;:") in unit_tokens:
            tokens.pop()
        return " ".join(tokens).strip(" -:;,.")

    @staticmethod
    def _looks_like_data_line(line: str) -> bool:
        text = normalize_description(line)
        if len(text) < 8:
            return False
        if any(token in text for token in ("item paquete", "sub item", "codigo sap", "cantidad por")):
            return False
        return bool(re.search(r"[a-zA-Z]", line) and re.search(r"\d", line))

    @staticmethod
    def _looks_like_weak_description(description: str) -> bool:
        normalized = normalize_description(description)
        if len(normalized) < 4:
            return True
        if re.search(r"\b\d{8}\b", description):
            return True
        return normalized in {"pba", "um", "descripcion"}

    @staticmethod
    def _has_useful_preview(rows: Iterable[ImportPreviewRow]) -> bool:
        return any(row.sap_code and row.quantity > 0 for row in rows)

    def _result(
        self,
        test: TenderTest,
        product: ReagentProduct,
        det_oc: float,
        det_env: float,
        quantity: int,
        warning: str = "",
    ) -> EquivalenceResult:
        return EquivalenceResult(
            test.sap_code,
            test.description,
            product.cod_prod,
            product.cod_eqv,
            product.product,
            product.det_rvo,
            det_oc,
            det_env,
            quantity,
            product.category or "Sin Categoría",
            warning,
        )

    @staticmethod
    def _looks_like_header(values) -> bool:
        text = normalize_description(" ".join(map(str, values)))
        return any(token in text for token in ("codigo sap", "descripcion", "cantidad", "codprod", "det rvo"))

    @staticmethod
    def _number(value: object, default: float = 0.0) -> float:
        if value is None:
            return default
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip().replace(" ", "")
        if not text:
            return default
        if "," in text and "." in text:
            text = text.replace(",", "")
        elif "," in text:
            if re.fullmatch(r"\d{1,3}(,\d{3})+", text):
                text = text.replace(",", "")
            else:
                text = text.replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return default

    @staticmethod
    def _text(value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return " ".join(str(value).strip().split())

    def _write_grouped_results(
        self,
        sheet,
        row: int,
        headers,
        results: list[EquivalenceResult],
        categories: Iterable[dict] | None = None,
    ) -> int:
        category_config = [
            item for item in (categories or ())
            if isinstance(item, dict) and str(item.get("name", "")).strip()
        ]
        colors = {
            str(item["name"]).strip(): str(item.get("color", "E2F0D9")).strip().lstrip("#")
            for item in category_config
        }
        category_order = {
            str(item["name"]).strip(): index
            for index, item in enumerate(category_config)
        }
        ordered_results = sorted(
            results,
            key=lambda item: (
                category_order.get(item.category.strip(), len(category_order)),
                item.category.casefold(),
                item.product.casefold(),
            ),
        )
        row = self._write_headers(sheet, row, headers)
        return self._write_result_rows(sheet, row, ordered_results, colors)

    @staticmethod
    def _write_result_rows(
        sheet,
        row: int,
        results: Iterable[EquivalenceResult],
        colors: dict[str, str] | None = None,
    ) -> int:
        for result in results:
            color = (colors or {}).get(result.category.strip(), "")
            fill = (
                PatternFill("solid", fgColor=color.upper())
                if re.fullmatch(r"[0-9A-Fa-f]{6}", color)
                and normalize_description(result.category) != "sin categoria"
                else None
            )
            values = (
                result.sap_code,
                result.test_description,
                result.cod_prod,
                result.cod_eqv,
                result.product,
                result.det_rvo if result.det_rvo > 0 else "",
                result.det_oc if result.det_oc > 0 else "",
                result.det_env if result.det_env > 0 else "",
                result.quantity if result.quantity > 0 else "",
            )
            for column, value in enumerate(values, 1):
                cell = sheet.cell(row, column, value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="left" if column in (2, 5) else "center")
                if column >= 6:
                    cell.number_format = "#,##0.##"
                if fill is not None and column == 5:
                    cell.fill = fill
            row += 1
        return row

    def _write_table(self, sheet, row: int, title: str, headers, rows, warning: bool = False) -> int:
        sheet.cell(row, 1, title)
        sheet.cell(row, 1).font = Font(bold=True, size=12)
        row += 1
        row = self._write_headers(sheet, row, headers, warning=warning)
        for values in rows:
            for column, value in enumerate(values, 1):
                cell = sheet.cell(row, column, value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center" if column != 2 else "left")
            row += 1
        return row

    @staticmethod
    def _write_headers(sheet, row: int, headers, warning: bool = False) -> int:
        for column, header in enumerate(headers, 1):
            cell = sheet.cell(row, column, header)
            cell.fill = WARNING_FILL if warning else HEADER_FILL
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
        return row + 1

    @staticmethod
    def _style_title(sheet, row: int, last_column: int) -> None:
        for column in range(1, last_column + 1):
            cell = sheet.cell(row, column)
            cell.fill = TITLE_FILL
            cell.font = Font(color="FFFFFF", bold=True, size=14)
            cell.alignment = Alignment(horizontal="center")
