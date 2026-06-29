from __future__ import annotations

import os
from datetime import date
from pathlib import Path
from unittest.mock import patch

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import load_workbook
from PySide6.QtWidgets import QApplication

from models.shipment import ShipmentAnalysis, ShipmentOptions, ShipmentRecord
from models.shipment_config import (
    ProductCategoryAssignment,
    ShipmentCategoryState,
    ShipmentForecastConfig,
    ShipmentForecastProductConfig,
    product_key,
)
from services.shipment_config_service import ShipmentConfigService
from services.shipment_forecast_service import (
    build_forecast_segments,
    forecast_date_range,
    parse_forecast_date,
)
from services.shipment_service import (
    FORECAST_DAY_COLUMN_WIDTH,
    FORECAST_START_COLUMN,
    ShipmentService,
)
from ui.dialogs.shipment_forecast_dialog import ShipmentForecastDialog
from ui.tabs.shipment_tab import ShipmentTab


def shipment(
    *,
    code: str = "P1",
    when: object = date(2024, 6, 15),
    quantity: float = 2,
    expires: object = "",
) -> ShipmentRecord:
    parsed = parse_forecast_date(when)
    return ShipmentRecord(
        fecha=when,  # type: ignore[arg-type]
        cliente="Cliente A",
        cliente_corto="CA",
        cod_prod=code,
        cod_eqv=f"E-{code}",
        producto=f"Producto {code}",
        cantidad=quantity,
        anio=parsed.year if parsed else 2024,
        mes=parsed.month if parsed else 6,
        linea="Línea A",
        categoria="Sin Categoría",
        expira=expires,  # type: ignore[arg-type]
    )


def configured(records: list[ShipmentRecord], *, enabled: bool = True, rate: float = 30) -> ShipmentForecastConfig:
    config = ShipmentForecastConfig(enabled=enabled, logistics_days=2)
    for record in records:
        key = product_key(record.cod_prod, record.cod_eqv, record.producto)
        config.products[key] = ShipmentForecastProductConfig(key, performance_days=rate)
    return config


def test_parse_fecha_ddmmyyyy_and_expira_with_spaces() -> None:
    assert parse_forecast_date("15/06/2024") == date(2024, 6, 15)
    assert parse_forecast_date("  20/07/2024  ") == date(2024, 7, 20)


def test_invalid_delivery_does_not_create_segment() -> None:
    record = shipment(when="fecha inválida")
    assert build_forecast_segments([record], configured([record])) == {}


def test_latest_delivery_is_used_and_same_day_quantities_are_summed() -> None:
    records = [
        shipment(when=date(2024, 5, 1), quantity=9),
        shipment(when=date(2024, 6, 15), quantity=2),
        shipment(when=date(2024, 6, 15), quantity=3),
    ]
    segment = next(iter(build_forecast_segments(records, configured(records, rate=1)).values()))
    assert segment.delivery_date == date(2024, 6, 15)
    assert segment.quantity == 5
    assert segment.performance_days == 5


def test_logistics_and_quantity_shift_performance() -> None:
    record = shipment(quantity=2)
    segment = next(iter(build_forecast_segments([record], configured([record])).values()))
    assert segment.performance_start == date(2024, 6, 17)
    assert segment.performance_days == 60
    assert segment.performance_end == date(2024, 8, 15)


def test_expiration_changes_performance_status_to_expired() -> None:
    record = shipment(quantity=1, expires=" 20/06/2024 ")
    segment = next(iter(build_forecast_segments([record], configured([record], rate=10)).values()))
    assert segment.status_on(date(2024, 6, 19)) == "performance"
    assert segment.status_on(date(2024, 6, 20)) == "expired"


def test_expiration_before_performance_marks_whole_performance_expired() -> None:
    record = shipment(quantity=1, expires="16/06/2024")
    segment = next(iter(build_forecast_segments([record], configured([record], rate=2)).values()))
    assert segment.status_on(segment.performance_start) == "expired"
    assert segment.status_on(segment.performance_end) == "expired"


def test_disabled_products_are_omitted_from_segments_and_range() -> None:
    first = shipment(code="P1", when=date(2024, 6, 1), quantity=1)
    second = shipment(code="P2", when=date(2024, 7, 1))
    config = configured([first, second], rate=10)
    config.products[product_key(second.cod_prod, second.cod_eqv, second.producto)].enabled = False
    segments = build_forecast_segments([first, second], config)
    assert len(segments) == 1
    assert forecast_date_range(segments.values()) == (date(2024, 6, 1), date(2024, 6, 12))


def test_calendar_range_uses_maximum_performance_end() -> None:
    first = shipment(code="P1", when=date(2024, 6, 1), quantity=1)
    second = shipment(code="P2", when=date(2024, 5, 1), quantity=14)
    config = configured([first, second], rate=30)
    config.products[product_key(second.cod_prod, second.cod_eqv, second.producto)].performance_days = 7
    segments = build_forecast_segments([first, second], config)
    assert forecast_date_range(segments.values()) == (
        date(2024, 5, 1),
        max(segment.performance_end for segment in segments.values()),
    )


def _analysis(records: list[ShipmentRecord]) -> ShipmentAnalysis:
    return ShipmentAnalysis(Path("input.xlsx"), tuple(records), len(records), 0)


def test_exporter_omits_calendar_when_disabled(tmp_path) -> None:
    record = shipment()
    output = ShipmentService(today=date(2024, 6, 20)).generate_report(
        _analysis([record]),
        tmp_path / "disabled.xlsx",
        ShipmentOptions(create_client_sheets=False),
        forecast_config=configured([record], enabled=False),
    )
    workbook = load_workbook(output)
    try:
        assert workbook["Total General"].cell(1, FORECAST_START_COLUMN).value is None
    finally:
        workbook.close()


def test_exporter_writes_calendar_headers_dimensions_and_colors(tmp_path) -> None:
    record = shipment(quantity=1, expires="20/06/2024")
    config = configured([record], rate=10)
    output = ShipmentService(today=date(2024, 6, 20)).generate_report(
        _analysis([record]),
        tmp_path / "forecast.xlsx",
        ShipmentOptions(create_client_sheets=False),
        forecast_config=config,
    )
    workbook = load_workbook(output)
    try:
        sheet = workbook["Total General"]
        assert sheet.cell(1, FORECAST_START_COLUMN).value == "JUNIO 2024"
        assert sheet.cell(2, FORECAST_START_COLUMN).value == "S"
        assert sheet.cell(3, FORECAST_START_COLUMN).value == 15
        assert sheet.row_dimensions[4].height == 15
        assert sheet.column_dimensions["T"].width == FORECAST_DAY_COLUMN_WIDTH
        assert sheet.cell(4, FORECAST_START_COLUMN).fill.fgColor.rgb.endswith(config.color("actual_delivery"))
    finally:
        workbook.close()


def test_forecast_configuration_round_trip_and_dialog_creation(tmp_path) -> None:
    key = product_key("P1", "E-P1", "Producto P1")
    state = ShipmentCategoryState(
        assignments={
            key: ProductCategoryAssignment(key, "P1", "E-P1", "Producto P1", "Sin Categoría", 0)
        }
    )
    state.forecast.enabled = True
    state.forecast.products[key] = ShipmentForecastProductConfig(key, False, 7, 3, "ABCDEF", "Urgencia")
    service = ShipmentConfigService(tmp_path / "shipment.json")
    service.save(state)
    loaded = service.load()
    assert loaded.forecast.enabled is True
    assert loaded.forecast.products[key].observation == "Urgencia"

    app = QApplication.instance() or QApplication([])
    dialog = ShipmentForecastDialog(loaded, service, {key})
    try:
        assert dialog.size().width() == 1200
        assert dialog.table.rowCount() == 1
    finally:
        dialog.close()
        app.processEvents()


def test_forecast_button_opens_dialog(tmp_path, monkeypatch) -> None:
    monkeypatch.setenv("AUTOMATIC_DATA_DIR", str(tmp_path))
    app = QApplication.instance() or QApplication([])
    tab = ShipmentTab(ShipmentService())
    tab.analysis = _analysis([shipment()])
    with patch("ui.tabs.shipment_tab.ShipmentForecastDialog") as dialog_class:
        tab.forecast_button.click()
        dialog_class.assert_called_once()
        dialog_class.return_value.exec.assert_called_once()
    tab.close()
    app.processEvents()
