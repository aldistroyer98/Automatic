from __future__ import annotations

from openpyxl import load_workbook

from app.paths import _migrate_legacy_data
from services.category_manager import CategoryManager
from services.equivalence_category_service import EquivalenceCategoryService
from services.equivalence_service import EquivalenceService
from services.excel_reader import sanitize_tabular_rows
from ui.window_sizes import (
    BASE_WINDOW_SIZE,
    DIALOG_1_2,
    DIALOG_1_3,
    DIALOG_1_4,
    DIALOG_2_3,
    DIALOG_3_4,
    DIALOG_5_6,
)


def test_standard_window_sizes_match_base_proportions() -> None:
    assert (BASE_WINDOW_SIZE.width(), BASE_WINDOW_SIZE.height()) == (1440, 810)
    assert (DIALOG_5_6.width(), DIALOG_5_6.height()) == (1200, 675)
    assert (DIALOG_3_4.width(), DIALOG_3_4.height()) == (1080, 608)
    assert (DIALOG_2_3.width(), DIALOG_2_3.height()) == (960, 540)
    assert (DIALOG_1_2.width(), DIALOG_1_2.height()) == (720, 405)
    assert (DIALOG_1_3.width(), DIALOG_1_3.height()) == (480, 270)
    assert (DIALOG_1_4.width(), DIALOG_1_4.height()) == (360, 203)


def test_sanitize_tabular_rows_removes_exported_index_and_empty_data() -> None:
    rows = [
        ("Unnamed: 0", "CodProd", "Producto", None),
        (0, "P-1", "Producto A", None),
        (None, None, None, None),
    ]

    assert sanitize_tabular_rows(rows) == [
        ("CodProd", "Producto"),
        ("P-1", "Producto A"),
    ]


def test_category_display_mapping_preserves_internal_value() -> None:
    assert CategoryManager.visible_name("Reactivo Principal") == "Producto Principal"
    assert CategoryManager.internal_name("Producto Principal") == "Reactivo Principal"


def test_legacy_data_migration_copies_without_deleting_source(tmp_path) -> None:
    source = tmp_path / "legacy"
    destination = tmp_path / "Automatic"
    source.mkdir()
    (source / "equivalence_config.json").write_text('{"version": 1}', encoding="utf-8")

    selected = _migrate_legacy_data(source, destination)

    assert selected == destination
    assert (destination / "equivalence_config.json").exists()
    assert (source / "equivalence_config.json").exists()


def test_product_import_ignores_unnamed_index_column(tmp_path) -> None:
    source = tmp_path / "products.xlsx"
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("Unnamed: 0", "CodProd", "CodEqv", "Producto", "DET RVO", "Categoría"))
    sheet.append((0, "P-1", "E-1", "Producto A", 5, "Sin Categoría"))
    workbook.save(source)

    products = EquivalenceService(tmp_path / "state.json").load_products_file(source)

    assert len(products) == 1
    assert products[0].cod_prod == "P-1"
    assert products[0].det_rvo == 5


def test_product_import_rejects_missing_required_headers(tmp_path) -> None:
    source = tmp_path / "invalid.xlsx"
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("CodProd", "Producto"))
    sheet.append(("P-1", "Producto A"))
    workbook.save(source)

    service = EquivalenceService(tmp_path / "state.json")
    try:
        service.load_products_file(source)
    except ValueError as exc:
        assert "Faltan encabezados obligatorios" in str(exc)
    else:
        raise AssertionError("Expected missing-header validation")


def test_product_import_rejects_non_numeric_det_rvo(tmp_path) -> None:
    from openpyxl import Workbook

    source = tmp_path / "invalid_det.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("CodProd", "CodEqv", "Producto", "DET RVO"))
    sheet.append(("P-1", "E-1", "Producto A", "no-numérico"))
    workbook.save(source)

    service = EquivalenceService(tmp_path / "state.json")
    try:
        service.load_products_file(source)
    except ValueError as exc:
        assert "DET RVO debe ser numérico" in str(exc)
    else:
        raise AssertionError("Expected DET RVO validation")


def test_product_import_allows_empty_det_rvo_and_category(tmp_path) -> None:
    from openpyxl import Workbook

    source = tmp_path / "empty_det.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("Orden", "CodProd", "CodEqv", "Producto", "DET RVO", "Categoría"))
    sheet.append((1, "P-1", "E-1", "Producto A", None, None))
    sheet.append((None, "", "E-2", "Producto B", "", ""))
    workbook.save(source)

    products = EquivalenceService(tmp_path / "state.json").load_products_file(source)

    assert [(product.det_rvo, product.category, product.order) for product in products] == [
        (None, "", 0),
        (None, "", 1),
    ]


def test_product_import_rejects_negative_det_rvo(tmp_path) -> None:
    from openpyxl import Workbook

    source = tmp_path / "negative_det.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("Orden", "CodProd", "CodEqv", "Producto", "DET RVO", "Categoría"))
    sheet.append((1, "P-1", "E-1", "Producto A", -1, "Consumible"))
    workbook.save(source)

    service = EquivalenceService(tmp_path / "state.json")
    try:
        service.load_products_file(source)
    except ValueError as exc:
        assert "DET RVO debe ser numérico y no negativo" in str(exc)
    else:
        raise AssertionError("Expected negative DET RVO validation")


def test_product_export_preserves_empty_det_rvo_cell(tmp_path) -> None:
    from models.equivalence import ReagentProduct

    service = EquivalenceService(tmp_path / "state.json")
    output = service.export_products_excel(
        tmp_path / "empty_det_export.xlsx",
        [ReagentProduct("P-1", "E-1", "Producto A", None, "", 0)],
    )

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        row = next(workbook.active.iter_rows(min_row=2, values_only=True))
    finally:
        workbook.close()

    assert row == (1, "P-1", "E-1", "Producto A", None, None)


def test_product_export_has_no_index_or_unnamed_columns(tmp_path) -> None:
    from models.equivalence import ReagentProduct

    service = EquivalenceService(tmp_path / "state.json")
    output = service.export_products_excel(
        tmp_path / "products.xlsx",
        [ReagentProduct("P-1", "E-1", "Producto A", 5, "Sin Categoría", 0)],
    )

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        headers = next(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()

    assert headers == ("Orden", "CodProd", "CodEqv", "Producto", "DET RVO", "Categoría")
    assert not any(str(value or "").startswith("Unnamed:") for value in headers)


def test_product_import_accepts_new_order_first_format(tmp_path) -> None:
    from openpyxl import Workbook

    source = tmp_path / "new_format.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("Orden", "CodProd", "CodEqv", "Producto", "DET RVO", "Categoría"))
    sheet.append((2, "P-2", "E-2", "Producto B", 20, "Reactivo Principal"))
    sheet.append((1, "P-1", "E-1", "Producto A", 10, "Reactivo Principal"))
    workbook.save(source)

    products = EquivalenceService(tmp_path / "state.json").load_products_file(source)

    assert [(product.cod_prod, product.order) for product in products] == [
        ("P-2", 1),
        ("P-1", 0),
    ]


def test_product_import_accepts_legacy_format_and_generates_missing_order(tmp_path) -> None:
    from openpyxl import Workbook

    service = EquivalenceService(tmp_path / "state.json")
    legacy = tmp_path / "legacy.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("CodProd", "CodEqv", "Producto", "DET RVO", "Categoría", "Orden"))
    sheet.append(("P-1", "E-1", "Producto A", 10, "Consumible", 3))
    workbook.save(legacy)
    assert service.load_products_file(legacy)[0].order == 2

    without_order = tmp_path / "without_order.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("CodProd", "CodEqv", "Producto", "DET RVO", "Categoría"))
    sheet.append(("P-1", "E-1", "Producto A", 10, "Consumible"))
    sheet.append(("P-2", "E-2", "Producto B", 20, "Consumible"))
    workbook.save(without_order)
    assert [product.order for product in service.load_products_file(without_order)] == [0, 1]


def test_product_export_round_trip_preserves_order(tmp_path) -> None:
    from models.equivalence import ReagentProduct

    service = EquivalenceService(tmp_path / "state.json")
    output = service.export_products_excel(
        tmp_path / "round_trip.xlsx",
        [
            ReagentProduct("P-2", "E-2", "Producto B", 20, "Consumible", 1),
            ReagentProduct("P-1", "E-1", "Producto A", 10, "Consumible", 0),
        ],
    )

    loaded = service.load_products_file(output)

    assert [(product.cod_prod, product.order) for product in loaded] == [
        ("P-1", 0),
        ("P-2", 1),
    ]


def test_equivalence_category_adapter_persists_separately(tmp_path) -> None:
    from models.equivalence import EquivalenceState, ReagentProduct
    from models.shipment_config import product_key

    persistence = EquivalenceService(tmp_path / "equivalence.json")
    state = EquivalenceState(
        products=[
            ReagentProduct("P-1", "E-1", "Producto A", None, "", 0),
            ReagentProduct("P-2", "E-2", "Producto B", 10, "", 1),
        ]
    )
    adapter = EquivalenceCategoryService(state, persistence)
    dialog_state = adapter.dialog_state()
    first = dialog_state.assignments[product_key("P-1", "E-1", "Producto A")]
    second = dialog_state.assignments[product_key("P-2", "E-2", "Producto B")]
    first.category_name = "Consumible"
    first.product_order = 1
    second.category_name = "Consumible"
    second.product_order = 0

    adapter.save(dialog_state)
    reloaded = persistence.load()

    assert [product.cod_prod for product in reloaded.products] == ["P-2", "P-1"]
    assert all(product.category == "Consumible" for product in reloaded.products)
    assert reloaded.products[1].det_rvo is None
    assert reloaded.settings["product_categories"]
