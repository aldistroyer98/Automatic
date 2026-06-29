from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from models.shipment import (
    ShipmentAnalysis,
    ShipmentFilterState,
    ShipmentOptions,
    ShipmentRecord,
)
from services.shipment_filters import (
    compute_available_filter_options,
    filter_records,
)
from services.shipment_service import INACTIVE_FILL, ShipmentService


def record(
    month: int,
    quantity: float,
    *,
    code: str = "P1",
    product: str = "Producto 1",
    client: str = "Cliente A",
    line: str = "Línea A",
    comodato: str = "C1",
    year: int = 2024,
) -> ShipmentRecord:
    return ShipmentRecord(
        fecha=date(year, month, 1),
        cliente=client,
        cliente_corto=client,
        cod_prod=code,
        cod_eqv=f"E-{code}",
        producto=product,
        cantidad=quantity,
        anio=year,
        mes=month,
        linea=line,
        categoria="Sin Categoría",
        comodato=comodato,
    )


def analysis(records: list[ShipmentRecord]) -> ShipmentAnalysis:
    return ShipmentAnalysis(
        source=Path("input.xlsx"),
        records=tuple(records),
        rows_read=len(records),
        ignored_rows=0,
    )


def test_two_empty_months_before_delivery_are_computable_zeroes() -> None:
    records = [
        record(4, 10, code="P1"),
        record(7, 5, code="P2", product="Producto 2"),
    ]
    rows = ShipmentService(today=date(2025, 1, 15)).preview(analysis(records))
    product = next(row for row in rows if row.cod_prod == "P1")

    assert product.meses == 4
    assert product.total == 10
    assert product.prod == 2.5


def test_three_empty_months_are_not_computable_and_later_delivery_reactivates() -> None:
    records = [record(3, 10), record(7, 6)]
    rows = ShipmentService(today=date(2025, 1, 15)).preview(analysis(records))

    assert rows[0].meses == 4
    assert rows[0].total == 16
    assert rows[0].prod == 4


def test_product_counts_zero_when_another_product_keeps_group_active() -> None:
    records = [
        record(4, 12, code="P1"),
        record(5, 3, code="P2", product="Producto 2"),
        record(6, 4, code="P2", product="Producto 2"),
    ]
    rows = ShipmentService(today=date(2025, 1, 15)).preview(analysis(records))
    product = next(row for row in rows if row.cod_prod == "P1")

    assert product.meses == 3
    assert product.total == 12
    assert product.prod == 4


def test_three_empty_group_months_are_blank_and_highlighted_in_excel(tmp_path) -> None:
    records = [record(3, 10), record(7, 6)]
    service = ShipmentService(today=date(2025, 1, 15))
    output = service.generate_report(
        analysis(records),
        tmp_path / "report.xlsx",
        ShipmentOptions(create_client_sheets=False),
    )

    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Total General"]
        assert sheet["G4"].value is None  # Abril
        assert sheet["H4"].value is None  # Mayo
        assert sheet["I4"].value is None  # Junio
        assert sheet["G4"].fill.fgColor.rgb == INACTIVE_FILL.fgColor.rgb
        assert sheet["Q4"].value == 4
        assert sheet["R4"].value == '=IF(Q4=0,"",P4/Q4)'
    finally:
        workbook.close()


def test_group_without_any_delivery_has_no_computable_months() -> None:
    records = [record(month, 0) for month in (5, 6, 7)]
    rows = ShipmentService(today=date(2025, 1, 15)).preview(analysis(records))

    assert rows[0].meses == 0
    assert rows[0].total == 0
    assert rows[0].prod is None


def test_january_and_february_before_march_delivery_are_computable() -> None:
    rows = ShipmentService(today=date(2025, 1, 15)).preview(
        analysis([record(3, 9)])
    )

    assert rows[0].meses == 3
    assert rows[0].total == 9
    assert rows[0].prod == 3


def test_current_month_is_not_excluded_when_today_is_june() -> None:
    service = ShipmentService(today=date(2024, 6, 15))

    assert service._last_valid_month(2024, ShipmentOptions(exclude_current_month=True)) == 6


def test_current_year_includes_current_month_in_divisor() -> None:
    rows = ShipmentService(today=date(2024, 6, 15)).preview(
        analysis([record(5, 10, year=2024)])
    )

    assert rows[0].meses == 2
    assert rows[0].total == 10
    assert rows[0].prod == 5


def test_previous_year_uses_december_as_final_month() -> None:
    rows = ShipmentService(today=date(2024, 6, 15)).preview(
        analysis([record(11, 8, year=2023)])
    )

    assert rows[0].meses == 2
    assert rows[0].total == 8
    assert rows[0].prod == 4


def test_future_year_only_uses_months_required_by_dataset() -> None:
    service = ShipmentService(today=date(2024, 6, 15))
    options = ShipmentOptions()

    assert service._last_valid_month(2025, options) == 0
    assert service._last_valid_month(2025, options, [record(2, 7, year=2025)]) == 2


def test_current_month_records_are_counted_in_total_months_and_average() -> None:
    rows = ShipmentService(today=date(2024, 6, 15)).preview(
        analysis([record(5, 10, year=2024), record(6, 5, year=2024)])
    )

    assert rows[0].total == 15
    assert rows[0].meses == 2
    assert rows[0].prod == 7.5


def test_excel_total_month_and_average_include_current_month(tmp_path) -> None:
    output = ShipmentService(today=date(2024, 6, 15)).generate_report(
        analysis([record(5, 10, year=2024), record(6, 5, year=2024)]),
        tmp_path / "current_month.xlsx",
        ShipmentOptions(create_client_sheets=False),
    )

    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Total General"]
        assert sheet["H4"].value is not None  # Mayo
        assert sheet["I4"].value is not None  # Junio
        assert sheet["P4"].value == "=SUM(D4:O4)"
        assert sheet["Q4"].value == 2
        assert sheet["R4"].value == '=IF(Q4=0,"",P4/Q4)'
    finally:
        workbook.close()


def test_three_empty_month_rule_evaluates_current_month() -> None:
    rows = ShipmentService(today=date(2024, 6, 15)).preview(
        analysis([record(3, 9, year=2024)])
    )

    assert rows[0].meses == 3
    assert rows[0].total == 9
    assert rows[0].prod == 3


def test_all_products_share_the_same_month_computability_in_excel(tmp_path) -> None:
    records = [
        record(3, 10, code="P1"),
        record(7, 6, code="P2", product="Producto 2"),
    ]
    output = ShipmentService(today=date(2025, 1, 15)).generate_report(
        analysis(records),
        tmp_path / "uniform_columns.xlsx",
        ShipmentOptions(create_client_sheets=False),
    )

    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook["Total General"]
        for coordinate in ("G4", "G5", "H4", "H5", "I4", "I5"):
            assert sheet[coordinate].value is None
            assert sheet[coordinate].fill.fgColor.rgb == INACTIVE_FILL.fgColor.rgb
        assert sheet["J4"].value is not None
        assert sheet["J5"].value is not None
        assert sheet["Q4"].value == sheet["Q5"].value == 4
    finally:
        workbook.close()


def test_filter_options_are_cross_filtered() -> None:
    records = [
        record(1, 1, client="A", line="L1", comodato="C1", year=2023),
        record(1, 1, client="B", line="L2", comodato="C2", year=2024),
        record(1, 1, client="C", line="L2", comodato="C3", year=2024),
    ]
    state = ShipmentFilterState(selected_years={2024})

    options = compute_available_filter_options(records, state)

    assert options["clients"] == ("B", "C")
    assert options["lines"] == ("L2",)
    assert options["comodatos"] == ("C2", "C3")
    assert {item.cliente for item in filter_records(records, state)} == {"B", "C"}
